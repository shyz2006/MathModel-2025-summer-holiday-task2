import pandas as pd
import gurobipy as gp
from gurobipy import GRB
import openpyxl

# 读取Excel文件中的地块面积数据
file1 = r"E:\存放桌面\数学建模资料汇总数学建模BOOM\2025暑期任务-2\C题\附件1.xlsx"  # 修改为附件1的实际文件路径
data1 = pd.read_excel(file1, sheet_name="乡村的现有耕地")
data2 = pd.read_excel(file1, sheet_name="乡村种植的农作物")
file2 = r"E:\存放桌面\数学建模资料汇总数学建模BOOM\2025暑期任务-2\C题\附件2.xlsx"  # 修改为附件1的实际文件路径
data3 = pd.read_excel(file2, sheet_name="2023年的农作物种植情况")
data4 = pd.read_excel(file2, sheet_name="2023年统计的相关数据")

# 已知数据(根据实际情况初始化)
T = 7
I = 2
J = 54
K = 41
p = 4
q = 5
M = 100000
S = data1["地块面积/亩"].tolist()
I_k = data2['Ik'].tolist()
Price = [
    [3.25, 7.5, 8.25, 7, 6.75, 3.5, 3, 6.75, 6, 7.5, 40, 1.5, 3.25, 5.5, 3.5, 7, 8, 6.75, 6.5, 3.75, 6.25, 5.5, 5.75, 5.25, 5.5, 6.5, 5, 5.75, 7, 5.25, 7.25, 4.5, 4.5, 4, 0, 0, 0, 0, 0, 0, 0],
    [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 9.6, 8.1, 7.8, 4.5, 7.5, 6.6, 6.9, 6.8, 6.6, 7.8, 6, 6.9, 8.4, 6.3, 8.7, 5.4, 5.4, 4.8, 2.5, 2.5, 3.25, 57.5, 19, 16, 100]
]
Request = [
    [57000, 21850, 22400, 33040, 9875, 170840, 132750, 71400, 30000, 12500, 1500, 35100, 36000, 14000, 10000, 21000, 36480, 26880, 6480, 30000, 35400, 43200, 0, 1800, 3600, 4050, 4500, 34400, 9000, 1500, 1200, 3600, 1800, 0, 0, 0, 0, 0, 0, 0, 0],
    [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 810, 2160, 900, 810, 0, 0, 0, 1080, 4050, 1350, 0, 0, 0, 1800, 150000, 100000, 36000, 9000, 7200, 18000, 4200]
]

df1 = pd.read_excel("cost.xlsx", sheet_name="第一季")
df2 = pd.read_excel("cost.xlsx", sheet_name="第二季")
Cost1 = df1.values.transpose()
Cost2 = df2.values.transpose()
Cost = [Cost1, Cost2]

df3 = pd.read_excel("Produce.xlsx", sheet_name="第一季")
df4 = pd.read_excel("Produce.xlsx", sheet_name="第二季")
Produce1 = df3.values.transpose()
Produce2 = df4.values.transpose()
Produce = [Produce1, Produce2]

# 创建模型
model = gp.Model("Crop_Planting")

# 决策变量
X = model.addVars(T, I, J, K, vtype=GRB.CONTINUOUS, name="X")
Y = model.addVars(T, I, J, K, vtype=GRB.BINARY, name="Y")
Z_rice = model.addVars(T, range(27, 35), vtype=GRB.BINARY, name="Z_Rice")
Z = model.addVars(T, I, K, vtype=GRB.CONTINUOUS, name="Z_Sold")
Z_excess = model.addVars(T, I, K, vtype=GRB.CONTINUOUS, name="Z_Excess")

# 定义目标函数
model.setObjective(
    gp.quicksum(
        Price[i][k] * Z[t, i, k] + 0.5 * Price[i][k] * Z_excess[t, i, k]
        - gp.quicksum(Cost[i][j][k] * X[t, i, j, k] for j in range(J))
        for t in range(T) for i in range(I) for k in range(K)
    ),
    GRB.MAXIMIZE
)

# 添加约束: Z_sold不能超过需求
model.addConstrs(
    # 用圆括号包裹生成器表达式
    (Z[t, i, k] <= Request[i][k]
    for t in range(T) for i in range(I) for k in range(K)
    ), name="Sold_Limit")

# 添加约束: Z_sold不能超过总产量
model.addConstrs(
    # 用圆括号包裹生成器表达式
    (Z[t, i, k] <= gp.quicksum(Produce[i][j][k] * X[t, i, j, k] for j in range(J))
    for t in range(T) for i in range(I) for k in range(K)
    ), name="Production_Limit_Sold")

# 添加约束: 超出部分Z_excess
model.addConstrs(
    # 用圆括号包裹生成器表达式
    (Z_excess[t, i, k] == gp.quicksum(Produce[i][j][k] * X[t, i, j, k] for j in range(J)) - Z[t, i, k]
    for t in range(T) for i in range(I) for k in range(K)
    ), name="Excess_Calculation")

# 约束3: 是否种植该作物
model.addConstrs(
    # 用圆括号包裹生成器表达式
    (X[t, i, j, k] <= M * Y[t, i, j, k]
    for t in range(T) for i in range(I) for j in range(J) for k in range(K)
    ), name="X_UpperBound_Y")

model.addConstrs(
    (X[t, i, j, k] >= 0.01 * Y[t, i, j, k]
    for t in range(T) for i in range(I) for j in range(J) for k in range(K)
), name="X_LowerBound_Y")

# 约束4: 每块地每季度种植面积总和不能超过地块总面积
for t in range(T):
    for i in range(I):
        for j in range(J):
            model.addConstr(
                gp.quicksum(X[t, i, j, k] for k in range(K)) <= S[j],
                name=f"Area_{t}_{i}_{j}"
            )

# 约束5: 三年内必须至少种植一次豆类作物
model.addConstrs(
    (gp.quicksum(X[t, i, j, k] * I_k[k] for t in range(2) for i in range(I) for k in range(K)) >= S[j]
    for j in range(J)
    ), name="Legume_First_Two_Years")

for j in range(J):
    for t in range(T - 2):
        model.addConstr(
            gp.quicksum(X[tt, i, j, k] * I_k[k] for tt in range(t, t + 3) for i in range(I) for k in range(K)) >= S[j],
            name=f"Legume_{j}_{t}"
        )

# 约束6: 同一种作物在同一片土地上不能连续两个季度种植
model.addConstrs(
    (X[t, i, j, k] * X[t, i + 1, j, k] <= S[j]
    for t in range(T) for j in range(J) for k in range(K) for i in range(I - 1)
    ), name="No_Consecutive_Planting")

model.addConstrs(
    (X[t, i + 1, j, k] * X[t + 1, i, j, k] <= S[j]
    for t in range(T - 1) for j in range(J) for k in range(K) for i in range(I - 1)
    ), name="No_Consecutive_Planting")

# 约束7: 最多种植p种作物
model.addConstrs(
    (gp.quicksum(Y[t, i, j, k] for k in range(K)) <= p
    for t in range(T) for i in range(I) for j in range(J)
    ), name="Max_Three_Crops")

# 添加约束: 每种作物最多种在q块地上
model.addConstrs(
    (gp.quicksum(Y[t, i, j, k] for j in range(J)) <= q
    for t in range(T) for i in range(I) for k in range(K)
    ), name="Max_Five_Plots_Per_Crop")

# 约束8: 确保粮食作物在连续年份的第一季不能连种
model.addConstrs(
    (X[t, 0, j, k] + X[t + 1, 0, j, k] <= S[j]
    for t in range(T - 1) for j in range(J) for k in range(1, 16)
    ), name="No_Consecutive_Years_For_Grain")

# 约束: 编号1-26的土地在第二季不种植任何作物
model.addConstrs(
    (X[t, 1, j, k] == 0
    for t in range(T) for j in range(26) for k in range(K)
    ), name="No_Planting_Second_Season_For_Lands_1_26")

# 约束: 编号为1-26的土地上只能种植编号为1-15的作物
model.addConstrs(
    (X[t, i, j, k] == 0
    for t in range(T) for i in range(I) for j in range(26) for k in range(15, 41)
    ), name="No_Planting_Crops_16_41_On_Lands_1_26")

# 约束: 编号为1-15的作物只能种植在编号为1-26的土地上
model.addConstrs(
    (X[t, i, j, k] == 0
    for t in range(T) for i in range(I) for j in range(26, J) for k in range(15)
    ), name="No_Planting_Crops_1_15_On_Lands_27_54")

# 约束: 编号为27-34的土地种植水稻
model.addConstrs(
    (gp.quicksum(X[t, i, j, k] for i in range(I) for k in range(K) if k == 15) <= M * Z_rice[t, j]
    for t in range(T) for j in range(27, 35)
    ), name="Rice_Planting_Only_Once")

# 确保水稻只能种植在单季
model.addConstrs(
    (gp.quicksum(X[t, i, j, 15] for i in range(I)) <= S[j]
    for t in range(T) for j in range(27, 35)
    ), name="Single_Season_Rice")

# 添加约束1: 如果种植了水稻,则第二季不种植任何作物
model.addConstrs(
    (gp.quicksum(X[t, i, j, k] for i in range(I) for k in range(K)) <= M * (1 - Z_rice[t, j])
    for t in range(T) for j in range(27, 35)
    ), name="No_Second_Season_If_Rice")

# 添加约束2: 第一季只能种植17-34号作物
model.addConstrs(
    (gp.quicksum(X[t, 0, j, k] for k in range(16, 35)) == gp.quicksum(X[t, 0, j, k] for k in range(16, 35))
    for t in range(T) for j in range(27, 35)
    ), name="First_Season_Crops_17_34")

# 添加约束3: 第二季只能种植35-37号作物
model.addConstrs(
    (gp.quicksum(X[t, 1, j, k] for k in range(34, 38)) == gp.quicksum(X[t, 1, j, k] for k in range(34, 38))
    for t in range(T) for j in range(27, 35)
    ), name="Second_Season_Crops_35_37")

# 添加约束: 编号为35-37的作物只能种植在编号为27-34的土地上
model.addConstrs(
    (X[t, i, j, k] == 0
    for t in range(T) for i in range(I) for j in range(26) for k in range(34, 37 + 1)
    ), name="No_Planting_Crops_35_37_On_Lands_1_26")

# 添加约束1: 编号为38-41的作物只能在35-50号地的第二季种植
model.addConstrs(
    (X[t, i, j, k] == 0
    for t in range(T) for j in range(35) for k in range(37, 41)
    ), name="No_Planting_Crops_38_41_On_Lands_1_34")

# 添加约束2: 编号为38-41的作物只能种植在第二季
model.addConstrs(
    (X[t, 0, j, k] == 0
    for t in range(T) for j in range(35, 51) for k in range(37, 41)
    ), name="No_Planting_Crops_38_41_First_Season")

# 设置相对Gap为0.5%(相对误差)
model.setParam('MIPGap', 0.01)

# 优化模型
model.optimize()

# 输出结果
if model.status == GRB.OPTIMAL:
    print(f"Optimal solution found with objective value: {model.objVal}")
    for t in range(T):
        for i in range(I):
            for j in range(J):
                for k in range(K):
                    if X[t, i, j, k].x > 0:
                        print(f"Year {t+1}, Season {i+1}, Land {j+1}, Crop {k+1}: {X[t, i, j, k].x} acres planted")
    print(f"Optimal solution found with objective value: {model.objVal} (元)")

# 加载Excel文件
file_path = "附件3/result1_1.xlsx"
wb = openpyxl.load_workbook(file_path)

# 选择每年的表格
for t in range(T):  # 遍历每一年
    ws = wb.worksheets[t]  # 选择第t+1张表
    for i in range(I):
        for j in range(J):
            for k in range(K):
                if X[t, i, j, k].x > 0:
                    if i == 0:
                        row = j + 2
                        column = k + 3
                        ws.cell(row=row, column=column, value=X[t, i, j, k].x)
                    else:
                        if j >= 26:
                            row = j + 30
                            column = k + 3
                            ws.cell(row=row, column=column, value=X[t, i, j, k].x)

# 保存修改后的文件
wb.save("附件3/result1_1.xlsx")
print("结果已成功写入7个表对应的Excel文件!")
print("结果已成功写入Excel文件!")